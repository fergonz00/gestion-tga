#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
extraer_lista.py — Saca precio de lista y costo de una lista de precios VW (Excel)
y (opcional) la concilia contra lo cargado en precios_lista (Supabase wjfgl).

De acá sale la solapa "Precios" de gestion.titogonzalez.online (tabla precios_lista),
que es la referencia de precios/costos para el resto de los paneles (motor, ventas).

--------------------------------------------------------------------------------
FORMATO del Excel VW (hoja "Lista Pesos"):
  - Encabezado ~fila 7:  "Nro.: 890   Vigencia : 03/06/2026"  -> lista_num
  - Datos desde la fila 13.
  - Columnas (1-based):
      B (2)  = COD  (6 dígitos, ej BZ31T4)  <- puede repetirse por variante MY/SE
      C (3)  = familia (Polo, Amarok, ...)
      D (4)  = descripción (Track MSI MT)
      E (5)  = MY
      AD (30)= PRECIO CONCESIONARIO · PREC.NETO   -> costo_concesionario (lo que TGA paga a VW)
      AH (34)= PRECIO SUGERIDO AL PÚBLICO · PREC.NETO -> precio_lista
  - La col INCENT viene en 0: los INCENTIVOS NO están en la lista (salen de la circular
    de Condiciones Comerciales -> tabla incentivos).
  - Ojo: los PREC.NETO traen decimales (ej 40240402,5). Al cargar se redondean a entero;
    contra el Sheet "Actual BT" pueden diferir ±1 peso por redondeo (inmaterial).

MAPEO a precios_lista: el campo `modelo` NO es la descripción del Excel sino el
`nombre_corto` del catálogo (catalogo_modelos), que es la clave con la que el motor y
la fórmula de Ventas buscan precio/costo. Un mismo COD puede dar varios modelos del
catálogo (ej AGDD8A -> Amarok Extreme / Hero = 92.427.900 y Black Style = 93.430.750;
son filas distintas del Excel con la misma COD pero distinto precio).

CÓMO SE CARGA una lista nueva (mes nuevo):
  1) python extraer_lista.py "Circular N°NN - Lista #NNN 6 digitos Excel.xlsx"
     -> imprime las filas y las deja en excel_<lista>.json
  2) Mapear cada fila del Excel al nombre_corto del catálogo (mismo criterio que junio).
  3) POST a la Web App:  accion=setprecioslista  con
     { mes:"YYYY-MM", listaNum, filas:[{modelo(nombre_corto), codigo, precioLista, costo}] }
     (guardarPreciosListaBulk valida contra catalogo_modelos: descarta modelos que nadie lee).
  4) Reconciliar con --db para confirmar (debe dar todo OK).

Uso:
  python extraer_lista.py "Circular N°72 - Lista #890 6 digitos Excel.xlsx"
  python extraer_lista.py "...xlsx" --db 2026-06   # además concilia contra Supabase
"""
import sys, os, json, re

def extraer(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Lista Pesos"] if "Lista Pesos" in wb.sheetnames else wb[wb.sheetnames[0]]
    # lista_num: buscar "Nro.: NNN" en las primeras filas
    lista_num = None
    for r in range(1, 12):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str):
                m = re.search(r"Nro\.?:\s*(\d+)", v)
                if m: lista_num = int(m.group(1))
    rows = []
    for r in range(13, ws.max_row + 1):
        cod = ws.cell(r, 2).value            # B
        if not cod or not str(cod).strip():
            continue
        costo = ws.cell(r, 30).value         # AD
        lista = ws.cell(r, 34).value         # AH
        if costo is None and lista is None:
            continue
        rows.append({
            "cod":   str(cod).strip(),
            "fam":   str(ws.cell(r, 3).value or "").strip(),   # C
            "desc":  str(ws.cell(r, 4).value or "").strip(),   # D
            "my":    ws.cell(r, 5).value,                      # E
            "costo": round(costo) if costo else None,
            "lista": round(lista) if lista else None,
        })
    return lista_num, rows

def conciliar(rows, mes):
    """Concilia el Excel contra precios_lista del mes en Supabase (wjfgl)."""
    import urllib.request
    # PAT desde C:\proyectos\.secrets\supabase.env
    env = os.path.join(os.path.dirname(__file__), "..", "..", ".secrets", "supabase.env")
    pat = None
    for line in open(env, encoding="utf8"):
        if line.startswith("SUPABASE_PAT=") and line.split("=", 1)[1].strip():
            pat = line.split("=", 1)[1].strip()
    ref = "wjfglsafgaltusmbnccl"
    q = ("select codigo, precio_lista::bigint as lista, costo_concesionario::bigint as costo "
         "from precios_lista where mes='%s'" % mes)
    req = urllib.request.Request(
        "https://api.supabase.com/v1/projects/%s/database/query" % ref,
        data=json.dumps({"query": q}).encode(),
        # User-Agent explícito: la Management API está detrás de Cloudflare y
        # bloquea (403 error 1010) el UA default de urllib.
        headers={"Authorization": "Bearer " + pat, "Content-Type": "application/json",
                 "User-Agent": "gestion-tga-extraer-lista/1.0"})
    db = json.load(urllib.request.urlopen(req))
    from collections import defaultdict
    ex = defaultdict(list)
    for x in rows:
        ex[x["cod"]].append((x["costo"], x["lista"]))
    ok, diffs = 0, []
    for r in db:
        pair = (int(r["costo"]), int(r["lista"]))
        if pair in ex.get(r["codigo"], []):
            ok += 1
        else:
            diffs.append("cod=%s DB(c=%s,l=%s) no matchea Excel%s" %
                         (r["codigo"], r["costo"], r["lista"], ex.get(r["codigo"])))
    print("\n=== Conciliación %s: %d/%d OK, %d diffs ===" % (mes, ok, len(db), len(diffs)))
    for d in diffs:
        print("  " + d)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    path = sys.argv[1]
    lista_num, rows = extraer(path)
    print("Lista #%s · %d filas" % (lista_num, len(rows)))
    for x in rows:
        print("%-8s %-10s %-34s MY%s  costo=%11s  lista=%11s" %
              (x["cod"], x["fam"], x["desc"][:34], x["my"], x["costo"], x["lista"]))
    out = os.path.join(os.path.dirname(os.path.abspath(path)), "excel_%s.json" % lista_num)
    json.dump({"lista_num": lista_num, "rows": rows}, open(out, "w", encoding="utf8"), ensure_ascii=False)
    print("-> %s" % out)
    if "--db" in sys.argv:
        conciliar(rows, sys.argv[sys.argv.index("--db") + 1])
